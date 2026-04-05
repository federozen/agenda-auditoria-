"""
engine.py — Lógica de auditoría de la agenda Olé.
Importable desde Streamlit o desde el notebook.
"""

import re, unicodedata, json, io, csv
from difflib import SequenceMatcher
from datetime import datetime, date, timedelta
from zoneinfo import ZoneInfo
import openpyxl

# Estos parsers convierten copy-paste directo del sitio de Ole a JSON interno.
# No necesitas consola ni archivos: ejecutas esta celda una sola vez al inicio.
import re
from datetime import date as _date

_MESES = {
    "enero":1,"febrero":2,"marzo":3,"abril":4,"mayo":5,"junio":6,
    "julio":7,"agosto":8,"septiembre":9,"octubre":10,"noviembre":11,"diciembre":12,
    "ene":1,"feb":2,"mar":3,"abr":4,"may":5,"jun":6,
    "jul":7,"ago":8,"sep":9,"oct":10,"nov":11,"dic":12,
}
_DIAS = {"lunes","martes","miercoles","miércoles","jueves","viernes","sabado","sábado","domingo"}
_CANAL_KW = ["ESPN","TNT","TyC","Fox Sports","DSports","DGO","Disney",
             "Basquet Pass","Básquet Pass","LPF PLAY","Apple TV","Telefe"]
_SKIP_SET = {"IMPERDIBLE","Formaciones y más datos","Más datos"}
_STATUS_SET = {"Finalizado","En vivo","Suspendido","Postergado","Por jugar"}
_STATUS_PRE = ("1º T","2º T","3º T","ET","Pen","PT","ST","HT","FT","Prorroga","Prórroga")

def _es_canal(s):
    s = s.strip()
    if not s or len(s) > 45: return False
    sl = s.lower()
    return any(k.lower() in sl for k in _CANAL_KW)

def _es_tiempo(s): return bool(re.match(r"^\d{1,2}:\d{2}$", s.strip()))
def _es_digito(s): return bool(re.match(r"^\d+$", s.strip()))
def _es_guion(s):  return s.strip() == "-"
def _es_score(s):  return bool(re.match(r"^\d+\s*-\s*\d+$", s.strip()))
def _es_fecha_inline(s): return bool(re.match(r"^\d{2}-\d{2}$", s.strip()))

def _es_junk(s):
    s = s.strip()
    if not s: return True
    if s in _SKIP_SET or s in _STATUS_SET: return True
    if any(s.startswith(p) for p in _STATUS_PRE): return True
    if _es_digito(s) or _es_guion(s) or _es_score(s): return True
    if _es_canal(s) or _es_tiempo(s): return True
    if re.match(r"^\d+[T'\u2019]\s*", s): return True
    if s.startswith("__") and s.endswith("__"): return True
    return False

def _es_fecha_header(s):
    return bool(re.search(r"\d{1,2}\s+de\s+(" + "|".join(_MESES) + r")\s+de\s+\d{4}", s.lower()))

def _parsear_fecha_header(s):
    m = re.search(r"(\d{1,2})\s+de\s+(" + "|".join(_MESES) + r")\s+de\s+(\d{4})", s.lower())
    if m: return _date(int(m.group(3)), _MESES[m.group(2)], int(m.group(1))).isoformat()
    return None

def _parsear_fecha_dia(s, anio=None):
    """Formato '3 abril' o '3 de abril'"""
    m = re.match(r"(\d{1,2})\s+(?:de\s+)?(" + "|".join(_MESES) + r")", s.strip().lower())
    if m:
        y = anio or _date.today().year
        return _date(y, _MESES[m.group(2)], int(m.group(1))).isoformat()
    return None

def _parsear_fecha_inline(s):
    """Formato '03-04'"""
    m = re.match(r"^(\d{2})-(\d{2})$", s.strip())
    if m:
        y = _date.today().year
        return _date(y, int(m.group(2)), int(m.group(1))).isoformat()
    return None

# ── PARSER AGENDA ─────────────────────────────────────────────────────────────
def parse_texto_agenda(texto):
    """
    Convierte copy-paste de la pagina Agenda de Ole a lista de eventos.
    Detecta automaticamente fechas, competiciones, partidos y canales.
    """
    lineas = texto.splitlines()
    eventos = []
    fecha_actual = None
    comp_actual = None
    i = 0
    n = len(lineas)

    while i < n:
        s = lineas[i].strip(); i += 1
        if not s: continue
        if _es_fecha_header(s):
            fecha_actual = _parsear_fecha_header(s); continue
        if _es_junk(s): continue
        if not fecha_actual: continue

        # Lookahead para clasificar: competicion vs equipo local
        lookahead = []
        j = i
        while j < n and len(lookahead) < 6:
            ns = lineas[j].strip()
            if ns: lookahead.append(ns)
            j += 1

        is_team = False
        if lookahead:
            la0 = lookahead[0]
            if la0 == s: is_team = True
            elif _es_tiempo(la0): is_team = True
            elif la0 in _STATUS_SET: is_team = True
            elif _es_canal(la0): is_team = True

        if not is_team:
            comp_actual = s; continue
        if not comp_actual: continue

        equipo_local = s
        # Saltar logo repeat
        if i < n and lineas[i].strip() == equipo_local: i += 1
        # Saltar canales pre-hora (bloque IMPERDIBLE)
        canales_pre = []
        while i < n and _es_canal(lineas[i].strip()):
            c = lineas[i].strip()
            if c not in canales_pre: canales_pre.append(c)
            i += 1
        # Consumir junk hasta hora o visitante
        tiempo = None
        while i < n:
            sl = lineas[i].strip()
            if not sl: i += 1; continue
            if _es_junk(sl) and not _es_tiempo(sl): i += 1; continue
            if _es_tiempo(sl): tiempo = sl; i += 1; break
            break
        # Saltar mas junk
        while i < n:
            sl = lineas[i].strip()
            if not sl or (_es_junk(sl) and not _es_canal(sl)): i += 1; continue
            break
        if i >= n: continue
        equipo_visitante = lineas[i].strip(); i += 1
        if _es_junk(equipo_visitante) or equipo_visitante == equipo_local: continue
        # Saltar logo repeat visitante
        if i < n and lineas[i].strip() == equipo_visitante: i += 1
        # Saltar "Formaciones..." / "Mas datos"
        if i < n and lineas[i].strip() in _SKIP_SET: i += 1
        # Recolectar canales
        canales = list(canales_pre)
        while i < n:
            sl = lineas[i].strip()
            if _es_canal(sl):
                if sl not in canales: canales.append(sl)
                i += 1
            elif not sl: i += 1
            else: break

        eventos.append({"date": fecha_actual, "time": tiempo,
            "name": f"{equipo_local} vs. {equipo_visitante}",
            "competition": comp_actual, "canales": canales})
    return eventos

# ── PARSER RESULTADOS ─────────────────────────────────────────────────────────
_SKIP_PAISES = {
    "Argentina","España","Francia","Italia","Alemania","Brasil","Ecuador",
    "USA","Uruguay","Colombia","Chile","Peru","Bolivia","Paraguay","Venezuela",
    "Inglaterra","Holanda","Portugal","Belgica","Suecia","Turquia","Escocia",
}

def parse_texto_resultados(texto, anio=None):
    """
    Convierte copy-paste de Ole Resultados (dia por dia) a lista de eventos.
    No incluye canales (Opta no los tiene).
    """
    if anio is None: anio = _date.today().year
    lineas = texto.splitlines()
    eventos = []
    fecha_actual = None
    comp_actual = None
    i = 0
    n = len(lineas)

    def skip_junk(s):
        if not s: return True
        if s in _SKIP_SET or s in _STATUS_SET or s in _SKIP_PAISES: return True
        if s.startswith("__") and s.endswith("__"): return True
        if any(s.startswith(p) for p in _STATUS_PRE): return True
        if _es_digito(s) or _es_guion(s) or _es_score(s): return True
        if re.match(r"^PT\s*\d+|^\d+['\u2019]", s): return True
        return False

    while i < n:
        s = lineas[i].strip(); i += 1
        if not s: continue
        # Dia de semana solo
        if s.lower() in _DIAS: continue
        # "3 abril" -> fecha
        f = _parsear_fecha_dia(s, anio)
        if f: fecha_actual = f; continue
        if skip_junk(s): continue
        if not fecha_actual: continue

        # Fecha inline "DD-MM" -> partido futuro
        if _es_fecha_inline(s):
            fecha_ev = _parsear_fecha_inline(s) or fecha_actual
            if i >= n: continue
            equipo_local = lineas[i].strip(); i += 1
            if not equipo_local or skip_junk(equipo_local): continue
            tiempo = None
            if i < n and _es_tiempo(lineas[i].strip()): tiempo = lineas[i].strip(); i += 1
            if i >= n: continue
            equipo_visitante = lineas[i].strip(); i += 1
            if not equipo_visitante or skip_junk(equipo_visitante): continue
            if comp_actual:
                eventos.append({"date": fecha_ev, "time": tiempo,
                    "name": f"{equipo_local} vs. {equipo_visitante}",
                    "competition": comp_actual, "canales": []})
            continue

        if _es_score(s) or _es_digito(s) or _es_guion(s) or _es_tiempo(s): continue
        if s in _STATUS_SET: continue

        # Lookahead para clasificar
        lookahead = []
        j = i
        while j < n and len(lookahead) < 4:
            ns = lineas[j].strip()
            if ns: lookahead.append(ns)
            j += 1

        is_team = False
        if lookahead:
            la0 = lookahead[0]
            if _es_digito(la0) or _es_guion(la0) or _es_score(la0): is_team = True
            elif _es_tiempo(la0): is_team = True

        if not is_team: comp_actual = s; continue
        if not comp_actual: continue

        equipo_local = s
        # Consumir score (digito - digito)
        while i < n:
            sl = lineas[i].strip()
            if not sl: i += 1; continue
            if _es_digito(sl) or _es_guion(sl) or _es_score(sl): i += 1; continue
            if _es_tiempo(sl): i += 1; continue
            break
        if i >= n: continue
        equipo_visitante = lineas[i].strip(); i += 1
        if skip_junk(equipo_visitante) or not equipo_visitante: continue
        # Saltar indicador de tiempo en vivo
        if i < n and re.match(r"^PT\s*\d+|^\d+['\u2019]", lineas[i].strip()): i += 1

        eventos.append({"date": fecha_actual, "time": None,
            "name": f"{equipo_local} vs. {equipo_visitante}",
            "competition": comp_actual, "canales": []})
    return eventos


# las celdas de paste puedan escribir antes de que se ejecute la celda de carga
archivos = {"ole_json": None, "opta_json": None, "directv": None, "flow": None, "lnb_json": None}


TOLERANCIA_HORARIO_MIN = 5

# Que trae SofaScore (mantener amplio)
COMPETICIONES_ACTIVAS = {
    "liga profesional argentina", "liga profesional",
    "copa de la liga profesional", "copa de la liga",
    "copa argentina", "supercopa argentina", "primera nacional",
    "torneo apertura", "torneo clausura",
    "conmebol libertadores", "libertadores", "copa libertadores",
    "conmebol sudamericana", "sudamericana", "copa sudamericana",
    "recopa sudamericana", "copa america",
    "eliminatorias", "world cup qualification", "world cup qualif",
    "international friendlies", "amistoso internacional",
    "fifa world cup", "nations league", "copa del mundo",
    "uefa champions league", "champions league",
    "uefa europa league", "europa league",
    "uefa conference league", "conference league",
    "laliga", "la liga", "laliga hypermotion",
    "premier league", "bundesliga", "serie a", "ligue 1",
    "fa cup", "copa del rey",
    "formula 1", "f1",
}

# Que se reporta como FALTANTE.
# Errores de horario y canal se reportan para CUALQUIER partido que ya este en Ole.
COMPETICIONES_INFORME = {
    # Futbol argentino
    "liga profesional", "liga profesional argentina",
    "copa de la liga", "copa de la liga profesional",
    "copa argentina", "supercopa argentina",
    # CONMEBOL
    "libertadores", "copa libertadores", "conmebol libertadores",
    "sudamericana", "copa sudamericana", "conmebol sudamericana",
    "recopa sudamericana", "copa america",
    "eliminatorias", "world cup qualification", "world cup qualif",
    # UEFA
    "champions league", "uefa champions league",
    # NOTA: "afc champions" NO esta aqui para no confundir con UEFA Champions
    "europa league", "uefa europa league",
    "conference league", "uefa conference league",
    # 5 grandes ligas
    "laliga", "la liga", "premier league", "bundesliga", "serie a", "ligue 1",
    # Selecciones masculinas
    "international friendlies", "amistoso internacional",
    "copa del mundo", "world cup", "nations league",
}

URGENTE_KEYWORDS = {
    "libertadores", "sudamericana", "champions", "world cup", "copa del mundo",
    "eliminatorias", "copa america", "copa argentina", "liga profesional",
    "copa de la liga", "supercopa", "recopa",
    "europa league", "conference league",
}

IMPORTANTE_KEYWORDS = {
    "laliga", "la liga", "premier league", "bundesliga", "serie a", "ligue 1",
    "international friendlies", "amistoso", "nations league",
}

print(f"  Competiciones para SofaScore: {len(COMPETICIONES_ACTIVAS)}")
print(f"  Competiciones que se informan como faltantes: {len(COMPETICIONES_INFORME)}")
print(f"  Tolerancia horaria: {TOLERANCIA_HORARIO_MIN} min")


import re, unicodedata, json
from difflib import SequenceMatcher
from datetime import datetime, timedelta, date
from zoneinfo import ZoneInfo

TZ_BA = ZoneInfo("America/Argentina/Buenos_Aires")

def norm_str(s):
    """Minusculas, sin tildes, sin puntuacion, colapsa espacios."""
    if not s: return ""
    s = unicodedata.normalize("NFD", str(s).lower())
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    traducciones = {
        "grand prix": "gp", "prix": "gp",
        "united states": "estados unidos",
        "brazil": "brasil", "japan": "japon",
        "germany": "alemania", "england": "inglaterra",
        "france": "francia", "italy": "italia",
        "spain": "espana",
        # Equipos comunes inglés -> español
        "atletico madrid": "atletico de madrid",
        "man city": "manchester city", "man utd": "manchester united",
        "man. city": "manchester city",
        "atl.": "atletico", "atl ": "atletico ",
        "b. dortmund": "borussia dortmund",
        "paris saint-germain": "psg", "paris sg": "psg",
        "paris saint germain": "psg",
        "sporting cp": "sporting lisboa", "sporting lisb": "sporting lisboa",
        "sp. lisboa": "sporting lisboa", "sp lisboa": "sporting lisboa",
        "inter milan": "inter", "inter miami cf": "inter miami",
        "ac milan": "milan", "as roma": "roma",
        "bayer 04": "bayer leverkusen", "bayer04": "bayer leverkusen",
        "rb leipzig": "rb leipzig",
        "gimnasia y esgrima mendoza": "gimnasia mendoza",
        "gimnasia y esgrima la plata": "gimnasia la plata",
        "central cordoba sde": "central cordoba",
        "central cordoba santiago del estero": "central cordoba",
        "club atletico union de santa fe": "union santa fe",
        "union de santa fe": "union santa fe",
        "ca talleres": "talleres", "ca lanus": "lanus",
        "ca independiente": "independiente",
        "ind. rivadavia": "independiente rivadavia",
        "sp. cristal": "sporting cristal",
        "dep. tolima": "deportes tolima",
        # Torneos: inglés SofaScore -> español Ole
        "argentina liga nacional": "liga nacional",
        "liga nacional de basquet": "liga nacional",
        "lnb argentina": "liga nacional",
        "liga profesional de futbol": "liga profesional",
        "primera b nacional": "primera nacional",
        "world cup qual": "eliminatorias",
        "world cup qualification": "eliminatorias",
        "knockout stage": "",
        "knockout phase": "",
        "group stage": "",
        "primera etapa": "",
        "apertura": "",
        "clausura": "",
    }
    for eng, esp in traducciones.items():
        s = s.replace(eng, esp)
    s = re.sub(r"[^a-z0-9 ]", " ", s)
    return re.sub(r"\s+", " ", s).strip()

STOPWORDS = {"de","la","el","los","las","club","atletico","ca","fc","cd","cf",
             "san","santa","rc","sc","ac","as","ss","us","if","bk","sk","fk",
             "real","sporting","united","city","the"}

def tokens(s):
    return {w for w in norm_str(s).split() if w not in STOPWORDS and len(w) > 2}

def similitud_nombres(a, b):
    """Retorna score 0.0-1.0 entre dos nombres de partidos."""
    def split_vs(s):
        return re.split(r"\s+vs\.?\s+", s.lower().replace(" vs.", " vs "), maxsplit=1)
    pa, pb = split_vs(a), split_vs(b)
    def score_equipo(t1, t2):
        n1, n2 = norm_str(t1), norm_str(t2)
        if not n1 or not n2: return 0.0
        if n1 == n2: return 1.0
        if f" {n1} " in f" {n2} " or f" {n2} " in f" {n1} ": return 0.95
        t1s, t2s = tokens(t1), tokens(t2)
        if t1s and t2s:
            inter = len(t1s & t2s)
            ov = max(inter / len(t1s), inter / len(t2s))
            if ov >= 0.6: return ov
        return SequenceMatcher(None, n1, n2).ratio()
    if len(pa) == 2 and len(pb) == 2:
        normal  = (score_equipo(pa[0], pb[0]) + score_equipo(pa[1], pb[1])) / 2
        cruzado = (score_equipo(pa[0], pb[1]) + score_equipo(pa[1], pb[0])) / 2
        return max(normal, cruzado)
    return score_equipo(a, b)

def norm_canal(s):
    """Normaliza nombre de canal: minúsculas, sin espacios ni puntuación."""
    return re.sub(r"[^a-z0-9]", "", str(s).lower())

# Marcas base de canales — para comparación tolerante
_MARCAS_CANAL = [
    "espn", "tnt", "tyc", "dsports", "dgo", "disney",
    "fox", "telefe", "tycplay", "lpfplay", "appletv",
    "basquetpass", "tntsports", "tntsportspremium",
]

def _marca_canal(s):
    """Extrae la marca base de un canal para comparación tolerante.
    'ESPN PREMIUM' -> 'espn', 'TNT Sports Premium' -> 'tnt', 'DGO' -> 'dgo'"""
    n = norm_canal(s)
    for marca in ["espnpremium","espn","tntsportspremium","tntsports","tnt",
                  "tycplay","tycsportsplay","tycsports","tyc",
                  "dsports","dgo","disney","foxsports","fox",
                  "basquetpass","lpfplay","appletv","telefe"]:
        if n.startswith(marca) or marca in n:
            return marca
    return n

def parsear_canales_ole(raw):
    """Acepta string, lista o string con separadores. Devuelve lista limpia."""
    if not raw: return []
    if isinstance(raw, list):
        resultado = []
        for item in raw: resultado.extend(parsear_canales_ole(item))
        return resultado
    s = str(raw).strip()
    if not s or s.lower() in ("none", "n/a", "-", ""): return []
    partes = re.split(r"[/,|]", s)
    return [p.strip() for p in partes if p.strip()]

def canales_coinciden(canales_ole, canal_fuente):
    """True si el canal de la fuente coincide con alguno de Ole.
    Tolerante: compara por marca base (ESPN vs ESPN PREMIUM = misma familia).
    Para reportar error de canal necesita ser marca DISTINTA.
    """
    if not canales_ole or not canal_fuente: return False
    nf = norm_canal(canal_fuente)
    mf = _marca_canal(canal_fuente)
    for c in canales_ole:
        no = norm_canal(c)
        mo = _marca_canal(c)
        if nf == no: return True          # match exacto
        if nf in no or no in nf: return True  # uno contiene al otro
        if mf == mo: return True          # misma familia (ESPN = ESPN PREMIUM)
    return False

def diff_minutos(h1, h2):
    try:
        a1, a2 = map(int, h1.split(":"))
        b1, b2 = map(int, h2.split(":"))
        return abs((a1*60+a2) - (b1*60+b2))
    except Exception: return None

def es_relevante(competicion_str):
    c = (competicion_str or "").lower()
    return any(kw in c for kw in COMPETICIONES_ACTIVAS)

def prioridad(competicion_str):
    c = (competicion_str or "").lower()
    if any(kw in c for kw in URGENTE_KEYWORDS): return "URGENTE"
    if any(kw in c for kw in IMPORTANTE_KEYWORDS): return "IMPORTANTE"
    return "REVISAR"

def limpiar_json(texto):
    texto = re.sub(r"^```[a-z]*\n?", "", texto.strip())
    return re.sub(r"```$", "", texto).strip()



# Cada funcion devuelve lista de dicts con esquema:
# {date, time, name, competition, canales:[], source}
import openpyxl, io

def _evento(date, time, name, competition, canales, source):
    return {
        "date":        date,
        "time":        time if time else None,
        "name":        (name or "").strip(),
        "competition": (competition or "").strip(),
        "canales":     canales if isinstance(canales, list) else parsear_canales_ole(canales),
        "source":      source,
    }

# 2a. JSON / CSV Ole agenda
# Acepta TRES formatos:
#   A) Scrapeado: lista con {fecha, hora, local, visitante, torneo, tv}
#      donde "fecha" es texto ("Hoy", "Domingo 5 de abril de 2026")
#   B) Estandar: lista/dict con {date, time, name, competition, canales}
#   C) CSV: mismo esquema que A pero en texto separado por comas
_MESES_OLE = {
    "enero":1,"febrero":2,"marzo":3,"abril":4,"mayo":5,"junio":6,
    "julio":7,"agosto":8,"septiembre":9,"octubre":10,"noviembre":11,"diciembre":12
}

def _parsear_fecha_texto(s):
    """Convierte 'Domingo 5 de abril de 2026' a ISO. Retorna None si no reconoce."""
    import re as _re
    m = _re.search(
        r"(\d{1,2})\s+de\s+(" + "|".join(_MESES_OLE) + r")\s+de\s+(\d{4})",
        s.lower()
    )
    if m:
        from datetime import date as _d2
        return _d2(int(m.group(3)), _MESES_OLE[m.group(2)], int(m.group(1))).isoformat()
    if re.match(r"\d{4}-\d{2}-\d{2}", s): return s[:10]
    return None

def _inferir_fecha_hoy(lista):
    """
    "Hoy" en Ole es el dia de exportacion.
    Lo inferimos como el dia anterior a la fecha mas baja con fecha explicita.
    Si no hay fechas explicitas, usamos date.today().
    """
    from datetime import date as _d2, timedelta
    fechas_explicitas = []
    for e in lista:
        s = str(e.get("fecha","")).strip()
        if s and s.lower() != "hoy":
            f = _parsear_fecha_texto(s)
            if f: fechas_explicitas.append(f)
    if fechas_explicitas:
        primera = min(fechas_explicitas)
        from datetime import date as _d3
        d = _d3.fromisoformat(primera)
        return (d - timedelta(days=1)).isoformat()
    return _d2.today().isoformat()

# _resolver_fecha_ole mantenida por compatibilidad — usa _parsear_fecha_texto
def _resolver_fecha_ole(fecha_str, *_):
    s = str(fecha_str).strip()
    if not s or s.lower() == "hoy": return None
    return _parsear_fecha_texto(s)

def _es_formato_scrapeado(item):
    """Detecta si un dict es del formato Ole scrapeado (tiene local/visitante/torneo)."""
    return "local" in item and "visitante" in item and "torneo" in item

def norm_ole_agenda(raw):
    from datetime import datetime as _dtnow
    from zoneinfo import ZoneInfo as _ZI
    texto = raw if isinstance(raw, str) else bytes(raw).decode("utf-8")

    # Detectar CSV
    primera_linea = texto.strip().splitlines()[0] if texto.strip() else ""
    es_csv = "," in primera_linea and not primera_linea.strip().startswith("{")

    if es_csv:
        import csv as _csv, io as _io
        texto = texto.lstrip("\ufeff")
        reader = _csv.DictReader(_io.StringIO(texto))
        parsed = []
        for row in reader:
            clean = {k.lstrip("\ufeff").strip(): v for k, v in row.items()}
            parsed.append(clean)
    else:
        parsed = json.loads(limpiar_json(texto))
        # Desenvolver wrapper
        if isinstance(parsed, dict):
            for clave in ("agenda","eventos","events","data","items"):
                if clave in parsed and isinstance(parsed[clave], list):
                    parsed = parsed[clave]; break
            else:
                for v in parsed.values():
                    if isinstance(v, list): parsed = v; break

    if not isinstance(parsed, list):
        raise ValueError(f"Formato no reconocido: {type(parsed)}")

    sin_fecha = 0
    eventos = []
    for e in parsed:
        if not isinstance(e, dict): continue

        if _es_formato_scrapeado(e):
            # ── Formato scrapeado Ole ──────────────────────────────────────
            # Descartar "Hoy" - solo auditar eventos con fecha futura explicita
            fecha_raw_str = str(e.get("fecha","")).strip()
            if not fecha_raw_str or fecha_raw_str.lower() == "hoy":
                continue
            fecha = _parsear_fecha_texto(fecha_raw_str)
            if not fecha: sin_fecha += 1; continue
            local = str(e.get("local","")).strip()
            visita = str(e.get("visitante","")).strip()
            if not local or not visita: continue
            name = f"{local} vs. {visita}"
            comp = str(e.get("torneo","")).strip().title()
            # TV: "ESPN / DGO / DISNEY + PREMIUM" -> lista
            # Split por " / " (con espacios) para no partir "DISNEY + PREMIUM"
            tv_raw = str(e.get("tv","")).strip()
            if tv_raw:
                # Primero intentar " / " con espacios
                if " / " in tv_raw:
                    canales = [c.strip() for c in tv_raw.split(" / ") if c.strip()]
                elif "/" in tv_raw:
                    canales = [c.strip() for c in tv_raw.split("/") if c.strip()]
                else:
                    canales = [tv_raw]
            else:
                canales = []
        else:
            # ── Formato estandar (date/name/competition) ───────────────────
            name = (e.get("name") or "").strip()
            if not name: continue
            fecha_raw = e.get("date") or e.get("fecha") or None
            if not fecha_raw: sin_fecha += 1; continue
            fecha = _parsear_fecha_texto(str(fecha_raw))
            if not fecha: sin_fecha += 1; continue
            comp = e.get("competition") or e.get("competicion") or e.get("torneo") or ""
            # Etiqueta limpia como competicion si disponible
            etiq = str(e.get("label") or e.get("etiqueta") or "").strip()
            if etiq:
                etiq_limpia = re.sub(r"[^\w\s\-\.\(\)/áéíóúñÁÉÍÓÚÑ]", "", etiq).strip()
                if etiq_limpia: comp = etiq_limpia
            canales = parsear_canales_ole(e.get("canales", e.get("canal", e.get("tv",""))))

        eventos.append(_evento(
            date=fecha,
            time=e.get("time") or e.get("hora") or None,
            name=name,
            competition=str(comp).strip(),
            canales=canales,
            source="ole_agenda",
        ))

    if sin_fecha:
        print(f"  Aviso: {sin_fecha} eventos descartados por fecha no reconocida")

    # Deduplicar: mismo partido + fecha + hora (o muy parecido) -> conservar el primero
    vistos = {}
    eventos_dedup = []
    for e in eventos:
        # Clave de deduplicacion: nombre normalizado + fecha
        clave = (norm_str(e["name"]), e["date"])
        if clave not in vistos:
            vistos[clave] = e
            eventos_dedup.append(e)
        else:
            # Si el duplicado tiene canal y el original no, actualizar canales
            orig = vistos[clave]
            if e["canales"] and not orig["canales"]:
                orig["canales"] = e["canales"]
    if len(eventos_dedup) < len(eventos):
        print(f"  Deduplicados: {len(eventos)-len(eventos_dedup)} eventos repetidos eliminados")
    return eventos_dedup

# 2b. JSON Opta / Resultados Ole (sin TV)
def norm_opta(raw):
    texto = raw if isinstance(raw, str) else bytes(raw).decode("utf-8")
    parsed = json.loads(limpiar_json(texto))
    if isinstance(parsed, dict):
        for clave in ("agenda","eventos","events","data","items","resultados"):
            if clave in parsed and isinstance(parsed[clave], list):
                parsed = parsed[clave]; break
        else:
            for v in parsed.values():
                if isinstance(v, list): parsed = v; break
    data = parsed
    eventos = []
    for e in data:
        if not isinstance(e, dict): continue
        comp = e.get("competition","")
        if not es_relevante(comp): continue
        name = (e.get("name") or "").strip()
        if not name: continue
        eventos.append(_evento(
            date=e.get("date",""), time=e.get("time"),
            name=name, competition=comp, canales=[], source="opta",
        ))
    return eventos

# 2c. Excel DIRECTV
def norm_directv(raw_bytes, fechas_validas=None):
    TIPOS_OK = {"futbol","basquet","basket","tenis","automovilismo","formula",
                "boxeo","rugby","mma","motociclismo","voley","beisbol","ciclismo","hockey"}
    wb = openpyxl.load_workbook(io.BytesIO(bytes(raw_bytes)), read_only=True)
    ws = wb.active
    eventos = []
    fecha_act = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0: continue
        cols = list(row) + [None]*12
        fecha_raw, _, hora_raw, tipo, torneo, local, _, visitante, canal = cols[:9]
        if isinstance(fecha_raw, datetime): fecha_act = fecha_raw.date().isoformat()
        if not fecha_act: continue
        if fechas_validas and fecha_act not in fechas_validas: continue
        tipo_str = norm_str(str(tipo) if tipo else "")
        if not any(t in tipo_str for t in TIPOS_OK): continue
        if "programa" in tipo_str: continue
        local_str = str(local).strip() if local else ""
        visit_str = str(visitante).strip() if visitante else ""
        if not local_str: continue
        name = f"{local_str} vs. {visit_str}" if visit_str else local_str
        m = re.search(r"(\d{1,2}:\d{2})", str(hora_raw) if hora_raw else "")
        hora = m.group(1) if m else None
        canal_str = str(canal).strip() if canal else ""
        if re.fullmatch(r"\d+", canal_str): canal_str = ""
        comp = str(torneo).strip() if torneo else ""
        if not es_relevante(comp): continue
        eventos.append(_evento(date=fecha_act, time=hora, name=name,
            competition=comp, canales=[canal_str] if canal_str else [], source="directv"))
    wb.close()
    return eventos

# 2d. Excel Flow
def norm_flow(raw_bytes, fechas_validas=None):
    DEPORTES_OK = {"futbol","basquet","tenis","automovilismo","formula",
                   "boxeo","rugby","mma","voley","beisbol","ciclismo","golf","motociclismo"}
    mapa_dia = {}
    if fechas_validas:
        fechas_dt = [datetime.fromisoformat(f).date() for f in fechas_validas]
        f_min = min(fechas_dt) - timedelta(days=3)
        f_max = max(fechas_dt) + timedelta(days=3)
        cur = f_min
        while cur <= f_max:
            mapa_dia.setdefault(cur.day, []).append(cur.isoformat())
            cur += timedelta(days=1)
    wb = openpyxl.load_workbook(io.BytesIO(bytes(raw_bytes)), read_only=True)
    ws = wb["Argentina"] if "Argentina" in wb.sheetnames else wb.active
    cache_fecha = {}
    eventos = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 2: continue
        cols = list(row) + [None]*6
        fecha_raw, hora_raw, deporte, evento_col, info, canal = cols[:6]
        if not deporte: continue
        dep_str = norm_str(str(deporte))
        if not any(d in dep_str for d in DEPORTES_OK): continue
        # Resolver fecha "Viernes 3" -> ISO usando mapa_dia anclado
        fr = str(fecha_raw).strip() if fecha_raw else ""
        if fr not in cache_fecha:
            m = re.search(r"(\d+)", fr)
            fa = None
            if m and fechas_validas:
                dia = int(m.group(1))
                opciones = mapa_dia.get(dia, [])
                if len(opciones) == 1:
                    fa = opciones[0]
                elif len(opciones) > 1:
                    fechas_dt_list = [datetime.fromisoformat(f).date() for f in fechas_validas]
                    centro = fechas_dt_list[len(fechas_dt_list)//2]
                    fa = min(opciones, key=lambda d: abs((datetime.fromisoformat(d).date()-centro).days))
            cache_fecha[fr] = fa
        fa_flow = cache_fecha[fr]
        if not fa_flow: continue
        if fechas_validas and fa_flow not in fechas_validas: continue
        if isinstance(hora_raw, datetime): hora = hora_raw.strftime("%H:%M")
        elif hasattr(hora_raw, "hour"): hora = f"{hora_raw.hour:02d}:{hora_raw.minute:02d}"
        else: hora = None
        info_str = str(info).strip() if info else ""
        evento_str = str(evento_col).strip() if evento_col else ""
        if not info_str and not evento_str: continue
        name = info_str if info_str else evento_str
        name = re.sub(r"\s+vs\.?\s+", " vs. ", name, flags=re.IGNORECASE)
        canal_str = str(canal).strip() if canal else ""
        if re.fullmatch(r"\d+", canal_str): canal_str = ""
        comp = evento_str
        if not es_relevante(comp): continue
        eventos.append(_evento(date=fa_flow, time=hora, name=name,
            competition=comp, canales=[canal_str] if canal_str else [], source="flow"))
    wb.close()
    return eventos

# 2e. JSON LNB
def norm_lnb(raw, fechas_validas=None):
    texto = raw if isinstance(raw, str) else bytes(raw).decode("utf-8")
    parsed = json.loads(limpiar_json(texto))
    if isinstance(parsed, dict):
        for clave in ("agenda","partidos","games","data","items"):
            if clave in parsed and isinstance(parsed[clave], list):
                parsed = parsed[clave]; break
        else:
            for v in parsed.values():
                if isinstance(v, list): parsed = v; break
    data = parsed
    eventos = []
    for e in data:
        if not isinstance(e, dict): continue
        if fechas_validas and e.get("date") not in fechas_validas: continue
        name = (e.get("name") or "").strip()
        if not name: continue
        eventos.append(_evento(
            date=e.get("date",""), time=e.get("time"),
            name=name,
            competition=e.get("competition","Liga Nacional de Basquet"),
            canales=parsear_canales_ole(e.get("canales", e.get("canal",""))),
            source="lnb",
        ))
    return eventos



UMBRAL_MATCH  = 0.72
UMBRAL_DUDOSO = 0.45

def encontrar_match(evento, candidatos, umbral=UMBRAL_MATCH):
    """Busca el mejor match para evento dentro de candidatos del mismo dia."""
    misma_fecha = [c for c in candidatos if c["date"] == evento["date"]]
    if not misma_fecha: return None, 0.0
    mejor, mejor_score = None, 0.0
    for c in misma_fecha:
        sc = similitud_nombres(evento["name"], c["name"])
        if sc > mejor_score:
            mejor_score = sc
            mejor = c
    if mejor_score >= umbral: return mejor, mejor_score
    return None, mejor_score

def validar_dudosos_ia(pares_dudosos, claude_client):
    """Valida en un solo batch los pares con score en zona gris."""
    if not pares_dudosos: return set()
    lista = "\n".join(
        f"{i+1}. OLE: '{p['ole']}' | FUENTE: '{p['fuente']}' (score: {p['score']:.2f})"
        for i, p in enumerate(pares_dudosos)
    )
    prompt = (
        "Sos experto en deportes de Argentina e internacionales.\n"
        "Determina si cada par refiere al MISMO evento deportivo.\n"
        "Considera idiomas distintos (Alemania=Germany), abreviaturas, variaciones de nombre.\n"
        "Responde SOLO con un JSON: lista de numeros de los pares que son el mismo evento.\n"
        "Ejemplo si los pares 1 y 3 coinciden: [1,3]  |  Si ninguno: []\n\n"
        f"Pares:\n{lista}"
    )
    try:
        r = claude_client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=400,
            messages=[{"role":"user","content":prompt}],
        )
        texto = r.content[0].text.strip()
        texto = re.sub(r"```[a-z]*\n?","",texto).replace("```","").strip()
        return set(json.loads(texto))
    except Exception as e:
        print(f"  Advertencia validacion IA: {e}")
        return set()




def _clave(nombre, fecha):
    return (norm_str(nombre), fecha)

def _dedup_faltantes(lista):
    ag = {}
    for e in lista:
        k = _clave(e["partido"], e["fecha"])
        if k not in ag:
            ag[k] = {**e, "_fuentes":[e["fuente"]], "_canales":list(e["canales_ref"])}
        else:
            if e["fuente"] not in ag[k]["_fuentes"]: ag[k]["_fuentes"].append(e["fuente"])
            for c in e["canales_ref"]:
                if c not in ag[k]["_canales"]: ag[k]["_canales"].append(c)
            if e["prioridad"]=="URGENTE": ag[k]["prioridad"]="URGENTE"
            elif e["prioridad"]=="IMPORTANTE" and ag[k]["prioridad"]=="REVISAR": ag[k]["prioridad"]="IMPORTANTE"
    resultado = []
    for e in ag.values():
        e["fuente"] = " + ".join(e["_fuentes"])
        e["canales_ref"] = e["_canales"]
        resultado.append(e)
    return sorted(resultado, key=lambda x:({"URGENTE":0,"IMPORTANTE":1,"REVISAR":2}[x["prioridad"]],x["fecha"],x["competicion"]))

def _dedup_horas(lista):
    ag = {}
    for e in lista:
        k = _clave(e["partido"], e["fecha"])
        if k not in ag: ag[k] = {**e, "_refs":{e["fuente"]:e["hora_fuente"]}}
        else:
            ag[k]["_refs"][e["fuente"]] = e["hora_fuente"]
            if e["diff_min"] > ag[k]["diff_min"]:
                ag[k]["diff_min"] = e["diff_min"]
                ag[k]["hora_fuente"] = e["hora_fuente"]
    resultado = []
    for e in ag.values():
        e["fuente"] = " + ".join(f"{f}={h}" for f,h in e["_refs"].items())
        # Es transmision solo si TODAS las fuentes difieren en 15 o 30 min
        e["es_transmision"] = e.get("es_transmision", False)
        resultado.append(e)
    return sorted(resultado, key=lambda x:-x["diff_min"])

def _dedup_canales(lista):
    ag = {}
    for e in lista:
        k = _clave(e["partido"], e["fecha"])
        if k not in ag: ag[k] = {**e, "_refs":[f"{e['fuente']}: {e['canal_ref']}"]}
        else:
            r = f"{e['fuente']}: {e['canal_ref']}"
            if r not in ag[k]["_refs"]: ag[k]["_refs"].append(r)
    resultado = []
    for e in ag.values():
        e["fuentes_canales"] = " | ".join(e["_refs"])
        resultado.append(e)
    return sorted(resultado, key=lambda x:({"URGENTE":0,"IMPORTANTE":1,"REVISAR":2}[x["prioridad"]],x["fecha"]))

def auditar(ole_eventos, fuentes, claude_client):
    hallazgos = {"faltantes_ole":[], "errores_horario":[], "canales_faltantes":[], "canales_incorrectos":[]}
    pares_dudosos = []
    matches_confirmados = {}

    # Paso A: buscar matches directos y acumular dudosos
    for nombre_fuente, eventos_fuente in fuentes.items():
        for ev_f in eventos_fuente:
            mejor_ole, score = encontrar_match(ev_f, ole_eventos)
            if mejor_ole and score >= UMBRAL_MATCH:
                matches_confirmados[id(ev_f)] = (mejor_ole, score)
            elif score >= UMBRAL_DUDOSO:
                pares_dudosos.append({
                    "ole":       mejor_ole["name"] if mejor_ole else "?",
                    "fuente":    ev_f["name"],
                    "score":     score,
                    "ev_fuente": ev_f,
                    "ev_ole_cand": mejor_ole,
                })

    # Paso B: validacion IA de dudosos (un solo batch)
    if pares_dudosos:
        print(f"  Validando {len(pares_dudosos)} pares dudosos con IA...", end=" ")
        confirmados_ia = validar_dudosos_ia(pares_dudosos, claude_client)
        print(f"confirmo {len(confirmados_ia)}")
        for idx, p in enumerate(pares_dudosos, 1):
            if idx in confirmados_ia and p["ev_ole_cand"]:
                matches_confirmados[id(p["ev_fuente"])] = (p["ev_ole_cand"], p["score"])

    # Paso C: clasificar hallazgos
    for nombre_fuente, eventos_fuente in fuentes.items():
        tiene_canal = nombre_fuente in ("directv","flow","lnb")
        for ev_f in eventos_fuente:
            ev_ole, score = matches_confirmados.get(id(ev_f), (None, 0.0))
            if ev_ole is None:
                comp_l = ev_f["competition"].lower()
                if any(kw in comp_l for kw in COMPETICIONES_INFORME):
                    hallazgos["faltantes_ole"].append({
                        "partido":     ev_f["name"],
                        "competicion": ev_f["competition"],
                        "fecha":       ev_f["date"],
                        "hora":        ev_f["time"],
                        "canales_ref": ev_f["canales"],
                        "fuente":      nombre_fuente,
                        "prioridad":   prioridad(ev_f["competition"]),
                    })
                continue
            # Error de horario
            # Regla: si Ole agenda Y Opta coinciden -> horario correcto, no reportar.
            # Si la diferencia es 15 o 30 min exactos -> puede ser transmision, avisar como AVISO.
            # Si es otra diferencia >= TOLERANCIA -> ERROR.
            if ev_f["time"] and ev_ole["time"]:
                dm = diff_minutos(ev_f["time"], ev_ole["time"])
                if dm is not None and dm >= TOLERANCIA_HORARIO_MIN:
                    es_transmision = dm in (15, 30)
                    hallazgos["errores_horario"].append({
                        "partido":       ev_ole["name"],
                        "competicion":   ev_ole["competition"] or ev_f["competition"],
                        "fecha":         ev_ole["date"],
                        "hora_ole":      ev_ole["time"],
                        "hora_fuente":   ev_f["time"],
                        "diff_min":      dm,
                        "fuente":        nombre_fuente,
                        "prioridad":     prioridad(ev_ole["competition"] or ev_f["competition"]),
                        "es_transmision": es_transmision,
                    })
            # Canal
            if tiene_canal and ev_f["canales"]:
                canal_ref = ev_f["canales"][0]
                canales_ole = ev_ole["canales"]
                if not canales_ole:
                    hallazgos["canales_faltantes"].append({
                        "partido":     ev_ole["name"],
                        "competicion": ev_ole["competition"],
                        "fecha":       ev_ole["date"],
                        "hora":        ev_ole["time"],
                        "canal_ref":   canal_ref,
                        "fuente":      nombre_fuente,
                        "prioridad":   prioridad(ev_ole["competition"]),
                    })
                elif not canales_coinciden(canales_ole, canal_ref):
                    hallazgos["canales_incorrectos"].append({
                        "partido":     ev_ole["name"],
                        "competicion": ev_ole["competition"],
                        "fecha":       ev_ole["date"],
                        "hora":        ev_ole["time"],
                        "canales_ole": canales_ole,
                        "canal_ref":   canal_ref,
                        "fuente":      nombre_fuente,
                        "prioridad":   prioridad(ev_ole["competition"]),
                    })

    # Paso D: deduplicar
    hallazgos["faltantes_ole"]       = _dedup_faltantes(hallazgos["faltantes_ole"])
    hallazgos["errores_horario"]     = _dedup_horas(hallazgos["errores_horario"])
    hallazgos["canales_faltantes"]   = _dedup_canales(hallazgos["canales_faltantes"])
    hallazgos["canales_incorrectos"] = _dedup_canales(hallazgos["canales_incorrectos"])
    return hallazgos




def ejecutar_auditoria(ole_raw, opta_raw, directv_raw, flow_raw, lnb_raw, cliente_anthropic):
    """
    Ejecuta la auditoría completa.
    Retorna dict con: ole_eventos, fuentes, hallazgos, horas_error, horas_aviso,
                      falt, s_can, e_can, fechas_a_auditar, fuentes_usadas, today_str
    """
    from datetime import datetime as _dt2
    from zoneinfo import ZoneInfo as _ZI2
    _HOY = date.today().isoformat()
    today_str = _dt2.now(_ZI2("America/Argentina/Buenos_Aires")).strftime("%d/%m/%Y %H:%M")

    if not ole_raw:
        raise ValueError("No se cargó la agenda Olé.")

    # Normalizar Ole
    ole_eventos = norm_ole_agenda(ole_raw)
    ole_eventos = [e for e in ole_eventos if e["date"] >= _HOY]
    fechas_ole  = sorted(set(e["date"] for e in ole_eventos))

    if not ole_eventos:
        raise ValueError("La agenda Olé no tiene eventos futuros (>= hoy).")

    fuentes = {}

    # Fixtures / Opta
    if opta_raw:
        todos_opta = [e for e in norm_opta(opta_raw) if e["date"] >= _HOY]
        fechas_opta = set(e["date"] for e in todos_opta)
        fechas_a_auditar = sorted(set(fechas_ole) | fechas_opta)
        fuentes["opta"] = [e for e in todos_opta if e["date"] in fechas_a_auditar]
    else:
        fechas_a_auditar = list(fechas_ole)

    # DIRECTV
    if directv_raw:
        ev = norm_directv(directv_raw, fechas_validas=set(fechas_a_auditar))
        if ev: fuentes["directv"] = ev

    # Flow
    if flow_raw:
        ev = norm_flow(flow_raw, fechas_validas=set(fechas_a_auditar))
        if ev: fuentes["flow"] = ev

    # LNB
    if lnb_raw:
        ev = norm_lnb(lnb_raw, fechas_validas=set(fechas_a_auditar))
        if ev: fuentes["lnb"] = ev

    if not fuentes:
        raise ValueError("No hay fuentes de referencia cargadas.")

    # Auditar
    hallazgos = auditar(ole_eventos, fuentes, cliente_anthropic)

    horas = hallazgos["errores_horario"]
    falt  = hallazgos["faltantes_ole"]
    s_can = hallazgos["canales_faltantes"]
    e_can = hallazgos["canales_incorrectos"]

    # Clasificar horarios
    opta_horas_ok = set()
    if "opta" in fuentes:
        for ev_o in fuentes["opta"]:
            opta_horas_ok.add((norm_str(ev_o["name"]), ev_o["date"], ev_o["time"]))

    horas_error, horas_aviso = [], []
    for e in horas:
        clave_ole = (norm_str(e["partido"]), e["fecha"], e["hora_ole"])
        if e["fuente"] not in ("opta",) and clave_ole in opta_horas_ok:
            continue
        if e.get("es_transmision"):
            horas_aviso.append(e)
        else:
            horas_error.append(e)

    fuentes_usadas = [k.upper() for k in fuentes.keys()]

    return {
        "ole_eventos":       ole_eventos,
        "fuentes":           fuentes,
        "hallazgos":         hallazgos,
        "horas_error":       horas_error,
        "horas_aviso":       horas_aviso,
        "falt":              falt,
        "s_can":             s_can,
        "e_can":             e_can,
        "fechas_a_auditar":  fechas_a_auditar,
        "fuentes_usadas":    fuentes_usadas,
        "today_str":         today_str,
    }


def generar_informe(resultado, cliente_anthropic):
    """
    Genera el informe IA a partir del resultado de ejecutar_auditoria().
    Retorna string con el memo en Markdown.
    """
    horas_error      = resultado["horas_error"]
    horas_aviso      = resultado["horas_aviso"]
    falt             = resultado["falt"]
    s_can            = resultado["s_can"]
    e_can            = resultado["e_can"]
    today_str        = resultado["today_str"]
    fechas_a_auditar = resultado["fechas_a_auditar"]
    fuentes_usadas   = resultado["fuentes_usadas"]

    _INFORME_KW = {
        "liga profesional", "copa de la liga", "copa argentina", "supercopa argentina",
        "libertadores", "copa libertadores", "conmebol libertadores",
        "sudamericana", "copa sudamericana", "conmebol sudamericana",
        "recopa sudamericana", "copa america",
        "eliminatorias", "world cup qualification", "world cup qualif",
        "champions league", "uefa champions league",
        "europa league", "uefa europa league",
        "conference league", "uefa conference league",
        "laliga", "la liga", "premier league", "bundesliga", "serie a", "ligue 1",
        "international friendlies", "amistoso internacional",
        "copa del mundo", "world cup", "nations league",
        "liga nacional",
    }

    def _rel(competicion):
        c = (competicion or "").lower()
        return any(kw in c for kw in _INFORME_KW)

    falt_ok  = [e for e in falt   if _rel(e.get("competicion",""))]
    horas_ok = [e for e in horas_error if _rel(e.get("competicion",""))]
    horas_av = [e for e in horas_aviso if _rel(e.get("competicion",""))]
    scan_ok  = [e for e in s_can  if _rel(e.get("competicion",""))]
    ecan_ok  = [e for e in e_can  if _rel(e.get("competicion",""))]

    total = sum(len(x) for x in [falt_ok, horas_ok, horas_av, scan_ok, ecan_ok])
    if total == 0:
        return "**Sin errores detectados en las competiciones auditadas.** La agenda está al día."

    def fmt_f(iso):
        try: return datetime.fromisoformat(iso).strftime("%d/%m")
        except: return iso[-5:]

    def _fmt_faltante(e):
        return {
            "fecha": fmt_f(e["fecha"]), "hora": e.get("hora") or "?",
            "partido": e["partido"], "competicion": e["competicion"],
            "canal_ref": ", ".join(e["canales_ref"]) if e.get("canales_ref") else "",
            "fuente": e["fuente"], "prioridad": e["prioridad"],
        }
    def _fmt_horario(e):
        return {
            "fecha": fmt_f(e["fecha"]), "partido": e["partido"],
            "competicion": e["competicion"],
            "hora_ole": e["hora_ole"], "hora_correcta": e["hora_fuente"],
            "diff_min": e["diff_min"], "fuente": e["fuente"],
            "prioridad": e["prioridad"],
            "posible_transmision": e.get("es_transmision", False),
        }
    def _fmt_canal_faltante(e):
        return {
            "fecha": fmt_f(e["fecha"]), "hora": e.get("hora") or "?",
            "partido": e["partido"], "competicion": e["competicion"],
            "canal_agregar": e.get("fuentes_canales",""),
        }
    def _fmt_canal_incorrecto(e):
        return {
            "fecha": fmt_f(e["fecha"]), "hora": e.get("hora") or "?",
            "partido": e["partido"], "competicion": e["competicion"],
            "canal_ole": ", ".join(e.get("canales_ole",[])),
            "canal_correcto": e.get("fuentes_canales",""),
        }

    falt_urg  = [e for e in falt_ok if e["prioridad"] == "URGENTE"]
    falt_imp  = [e for e in falt_ok if e["prioridad"] == "IMPORTANTE"]
    horas_urg = [e for e in horas_ok if e["prioridad"] == "URGENTE"]

    resumen = (
        f"Fecha: {today_str} | Fechas auditadas: {', '.join(fechas_a_auditar)}\n"
        f"Fuentes: {', '.join(fuentes_usadas)}\n\n"
        f"TOTALES:\n"
        f"  Faltantes urgentes: {len(falt_urg)}\n"
        f"  Faltantes importantes: {len(falt_imp)}\n"
        f"  Horarios incorrectos: {len(horas_ok)} ({len(horas_urg)} urgentes)\n"
        f"  Avisos horario (15/30 min): {len(horas_av)}\n"
        f"  Canales faltantes: {len(scan_ok)}\n"
        f"  Canales incorrectos: {len(ecan_ok)}\n"
    )

    datos = {
        "fecha_auditoria": today_str,
        "fechas_auditadas": fechas_a_auditar,
        "fuentes": fuentes_usadas,
        "faltantes": [_fmt_faltante(e) for e in falt_ok],
        "horarios_incorrectos": [_fmt_horario(e) for e in horas_ok],
        "avisos_transmision": [_fmt_horario(e) for e in horas_av],
        "canales_faltantes": [_fmt_canal_faltante(e) for e in scan_ok],
        "canales_incorrectos": [_fmt_canal_incorrecto(e) for e in ecan_ok],
    }

    prompt = (
        "Sos el editor de la agenda deportiva del diario Ole.\n"
        "Redacta un memo interno en Markdown para el equipo de carga.\n"
        "El memo debe ser claro, concreto y accionable.\n\n"
        "USA ESTE FORMATO EXACTO (con Markdown):\n\n"
        "## Resumen\n"
        "[Una linea con totales de cada tipo de error]\n\n"
        "## Partidos faltantes\n"
        "- **[DD/MM HH:MM]** [Partido] — *[Competicion]* — canal: [X] *(fuente: Y)*\n"
        "(una linea por partido, ordenados por fecha)\n\n"
        "## Horarios a corregir\n"
        "- **[Partido]** [DD/MM] — tiene [HH:MM], debe ser **[HH:MM]** · diff [X]min *(fuente: Y)*\n"
        "(si posible_transmision=true agregar: *· verificar: puede ser inicio de transmision*)\n\n"
        "## Avisos de horario\n"
        "- **[Partido]** [DD/MM] — Ole: [HH:MM] / grilla: [HH:MM] *· puede ser inicio de transmision*\n\n"
        "## Canales a agregar\n"
        "- **[DD/MM HH:MM]** [Partido] — agregar **[canal]**\n\n"
        "## Canales a corregir\n"
        "- **[DD/MM HH:MM]** [Partido] — tiene **[canal_ole]**, debe ser **[canal_correcto]**\n\n"
        "## Prioridad maxima\n"
        "1. [accion concreta]\n"
        "(hasta 5 items, los mas urgentes primero)\n\n"
        "REGLAS:\n"
        "- Incluir TODOS los items de los datos. No omitir ninguno.\n"
        "- Omitir secciones que no tengan datos.\n"
        "- No agregar texto fuera del formato.\n\n"
        f"DATOS:\n{json.dumps(datos, ensure_ascii=False, indent=2)}"
    )

    try:
        r = cliente_anthropic.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=8096,
            messages=[{"role": "user", "content": prompt}],
        )
        return r.content[0].text.strip()
    except Exception as ex:
        return f"Error al generar informe: {ex}"


def resumen_por_dia(resultado):
    """
    Retorna lista de dicts por fecha. Cada fila tiene campos explícitos
    para hora, horario de referencia, canal Ole, canal referencia y notas.
    """
    ole_eventos      = resultado["ole_eventos"]
    horas_error      = resultado["horas_error"]
    horas_aviso      = resultado["horas_aviso"]
    falt             = resultado["falt"]
    s_can            = resultado["s_can"]
    e_can            = resultado["e_can"]
    fechas_a_auditar = resultado["fechas_a_auditar"]

    def _key(nombre, fecha):
        return (norm_str(nombre), fecha)

    idx_horas_error = {_key(e["partido"], e["fecha"]): e for e in horas_error}
    idx_horas_aviso = {_key(e["partido"], e["fecha"]): e for e in horas_aviso}
    idx_scan        = {_key(e["partido"], e["fecha"]): e for e in s_can}
    idx_ecan        = {_key(e["partido"], e["fecha"]): e for e in e_can}

    falt_por_fecha = {}
    for e in falt:
        falt_por_fecha.setdefault(e["fecha"], []).append(e)

    dias = []
    for fecha in sorted(fechas_a_auditar):
        filas = []

        # ── Eventos que SÍ están en Olé ──────────────────────────────────────
        for ev in [e for e in ole_eventos if e["date"] == fecha]:
            k = _key(ev["name"], fecha)

            hora_ole  = ev["time"] or "?"
            hora_ref  = ""          # hora según fuente externa
            hora_ok   = True

            canal_ole = ", ".join(ev["canales"]) if ev["canales"] else ""
            canal_ref = ""          # canal según fuente externa
            canal_ok  = True

            notas = []
            estado = "ok"

            # Horario incorrecto
            if k in idx_horas_error:
                h = idx_horas_error[k]
                hora_ref = h["hora_fuente"]
                hora_ok  = False
                estado   = "error"
                notas.append(f"fuente: {h['fuente']}")

            # Aviso de horario (posible transmisión)
            elif k in idx_horas_aviso:
                h = idx_horas_aviso[k]
                hora_ref = h["hora_fuente"]
                hora_ok  = None   # None = aviso, no error claro
                if estado == "ok": estado = "aviso"
                notas.append(f"posible inicio transmisión ({h['diff_min']}min, {h['fuente']})")

            # Canal faltante (Olé no tiene canal)
            if k in idx_scan:
                c = idx_scan[k]
                canal_ref = c.get("fuentes_canales", "")
                canal_ok  = None   # None = faltante
                if estado == "ok": estado = "aviso"
                notas.append(f"fuente: {c.get('fuente','')}")

            # Canal incorrecto
            elif k in idx_ecan:
                c = idx_ecan[k]
                canal_ref = c.get("fuentes_canales", "")
                canal_ok  = False
                if estado in ("ok", "aviso"): estado = "error"
                notas.append(f"fuente: {c.get('fuente','')}")

            filas.append({
                "en_ole":    True,
                "estado":    estado,
                "hora_ole":  hora_ole,
                "hora_ref":  hora_ref,
                "hora_ok":   hora_ok,        # True=ok, False=error, None=aviso
                "partido":   ev["name"],
                "competicion": ev["competition"],
                "canal_ole": canal_ole,
                "canal_ref": canal_ref,
                "canal_ok":  canal_ok,       # True=ok, False=error, None=faltante
                "notas":     " · ".join(notas),
            })

        # ── Faltantes en Olé ──────────────────────────────────────────────────
        for e in falt_por_fecha.get(fecha, []):
            canal_ref = ", ".join(e.get("canales_ref", [])) if e.get("canales_ref") else ""
            filas.append({
                "en_ole":    False,
                "estado":    "faltante",
                "hora_ole":  e.get("hora") or "?",
                "hora_ref":  e.get("hora") or "?",
                "hora_ok":   True,
                "partido":   e["partido"],
                "competicion": e["competicion"],
                "canal_ole": "",
                "canal_ref": canal_ref,
                "canal_ok":  None,
                "notas":     f"fuente: {e['fuente']}",
            })

        filas.sort(key=lambda x: x["hora_ole"])
        dias.append({"fecha": fecha, "filas": filas})
    return dias


# ── Funciones auxiliares expuestas para la app ────────────────────────────────

def _next_data(html):
    """Extrae __NEXT_DATA__ de HTML de Next.js."""
    import re as _re2
    m = _re2.search(r'<script id="__NEXT_DATA__"[^>]*>(.+?)</script>', html, re.DOTALL)
    try: return json.loads(m.group(1)) if m else None
    except: return None

def _buscar_evs(obj, acc=None, d=0):
    """Busca recursivamente eventos en JSON de Next.js."""
    if acc is None: acc = []
    if d > 12: return acc
    if isinstance(obj, list):
        for x in obj: _buscar_evs(x, acc, d+1)
    elif isinstance(obj, dict):
        ks = set(obj)
        if ((("homeTeam" in ks or "home" in ks) and ("awayTeam" in ks or "away" in ks))
                and ("date" in ks or "fecha" in ks or "startDate" in ks)):
            acc.append(obj)
        else:
            for v in obj.values(): _buscar_evs(v, acc, d+1)
    return acc

def _leer_xlsx_fixtures(raw_bytes):
    """Lee Excel de fixtures (hoja RESUMEN) y devuelve lista de eventos estándar."""
    wb = openpyxl.load_workbook(io.BytesIO(bytes(raw_bytes)), read_only=True)
    ws = wb["RESUMEN"] if "RESUMEN" in wb.sheetnames else wb.active
    header = None
    eventos = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            header = [str(c).strip().lower() if c else "" for c in row]
            continue
        if not any(row): continue
        def col(*names):
            for name in names:
                for h in header:
                    if name in h:
                        idx = header.index(h)
                        return row[idx] if idx < len(row) else None
            return None
        fecha_raw = col("fecha","date")
        hora_raw  = col("hora","time")
        ev_raw    = col("evento","event","partido","match","name")
        comp_raw  = col("competencia","competition","liga","league")
        etiq_raw  = col("etiqueta","label","tag")
        if isinstance(fecha_raw, datetime): fecha = fecha_raw.date().isoformat()
        elif fecha_raw: fecha = str(fecha_raw)[:10]
        else: continue
        if isinstance(hora_raw, datetime): hora = hora_raw.strftime("%H:%M")
        elif hasattr(hora_raw,"hour"): hora = f"{hora_raw.hour:02d}:{hora_raw.minute:02d}"
        elif hora_raw:
            m3 = re.search(r"\d{1,2}:\d{2}", str(hora_raw))
            hora = m3.group(0) if m3 else None
        else: hora = None
        name = str(ev_raw).strip() if ev_raw else ""
        if not name: continue
        comp = str(comp_raw).strip() if comp_raw else ""
        etiq = str(etiq_raw).strip() if etiq_raw else ""
        etiq_limpia = re.sub(r"[^\w\s\-\.\(\)/áéíóúñÁÉÍÓÚÑ]", "", etiq).strip()
        competicion = etiq_limpia if etiq_limpia else comp
        eventos.append({"date":fecha,"time":hora,"name":name,
                        "competition":competicion,"canales":[]})
    wb.close()
    return eventos
